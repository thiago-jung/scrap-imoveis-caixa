"""
Scraper diário de imóveis em leilão/venda da Caixa Econômica Federal.

Fonte: https://venda-imoveis.caixa.gov.br/listaweb/Lista_imoveis_{UF}.csv
(mesmo arquivo gerado pelo botão "Baixe a lista completa de imóveis" do site oficial)

O script:
  1. Baixa o CSV de cada UF configurada
  2. Filtra por cidade (default: Região Metropolitana de Porto Alegre)
  3. Salva um snapshot datado em disco (histórico)
  4. Compara com o snapshot anterior (novos/removidos)
  5. Gera um relatório HTML estático (index.html) pra abrir no navegador

Uso:
    python caixa_leiloes_scraper.py --ufs RS SP --outdir ./data

Agendamento sugerido:
    - cron:            0 7 * * *  python /path/caixa_leiloes_scraper.py --ufs RS
    - GitHub Actions:  schedule com cron diário (ver exemplo no final do arquivo)
"""

import argparse
import csv
import io
import logging
import os
import unicodedata
import webbrowser
from datetime import datetime
from pathlib import Path

import requests
from scrapers.mega_leiloes import MegaLeiloesScraper
from scrapers.leiloes_judiciais import LeiloesJudiciaisScraper

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

BASE_URL = "https://venda-imoveis.caixa.gov.br/listaweb/Lista_imoveis_{uf}.csv"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    )
}

# Região Metropolitana de Porto Alegre (principais cidades).
# Edite essa lista à vontade -- é só o filtro padrão.
CIDADES_RMPA = [
    "PORTO ALEGRE",
    "CANOAS",
    "ALVORADA",
    "VIAMAO",
    "CACHOEIRINHA",
    "GRAVATAI",
    "GUAIBA",
    "ESTEIO",
    "SAPUCAIA DO SUL",
    "NOVO HAMBURGO",
    "SAO LEOPOLDO",
    "ELDORADO DO SUL",
]


def normalizar(texto: str) -> str:
    """Remove acentos e deixa em maiúsculas, pra comparar cidade sem depender de encoding."""
    nfkd = unicodedata.normalize("NFKD", texto)
    sem_acento = "".join(c for c in nfkd if not unicodedata.combining(c))
    return sem_acento.upper()


# ---------------------------------------------------------------------------
# Índice de risco por município (criminalidade, normalizada por população)
#
# Fonte: SSP-RS via portal de dados abertos do RS (dados.rs.gov.br), datasets
# "Indicadores Criminais de <ano>" -- ocorrências por município, publicadas
# mensalmente sob a Lei Estadual 15.610/2021.
#
# LIMITAÇÃO IMPORTANTE: a granularidade oficial e programaticamente acessível
# é por MUNICÍPIO, não por bairro. Isso significa que todos os imóveis dentro
# de Porto Alegre recebem o mesmo nível de risco, mesmo sabendo que a
# variação de segurança entre bairros de POA é grande. É um proxy imperfeito,
# não um índice de risco por endereço.
#
# Criminalidade não muda dia a dia -- por isso esses dados são cacheados em
# disco e só re-baixados a cada CACHE_RISCO_DIAS dias, em vez de toda execução.
# ---------------------------------------------------------------------------

CKAN_API_INDICADORES = "https://dados.rs.gov.br/api/3/action/package_show"
CACHE_RISCO_DIAS = 25

# População (Censo IBGE 2022) dos municípios da Região Metropolitana de Porto
# Alegre cobertos pelo filtro CIDADES_RMPA. Atualize se a lista de cidades mudar.
POPULACAO_MUNICIPIOS_RMPA = {
    "PORTO ALEGRE": 1_332_570,
    "CANOAS": 347_657,
    "ALVORADA": 187_315,
    "VIAMAO": 224_116,
    "CACHOEIRINHA": 136_258,
    "GRAVATAI": 265_070,
    "GUAIBA": 92_924,
    "ESTEIO": 76_137,
    "SAPUCAIA DO SUL": 132_107,
    "NOVO HAMBURGO": 227_732,
    "SAO LEOPOLDO": 217_409,
    "ELDORADO DO SUL": 39_559,
}


def _consultar_recursos_indicadores_criminais(anos=(2026, 2025)) -> list[dict]:
    """Pergunta pra API do dados.rs.gov.br quais arquivos XLSX de indicadores
    criminais existem pros anos informados. Usa a API (não links fixos) pra
    não quebrar quando novos meses/anos forem publicados."""
    recursos = []
    for ano in anos:
        try:
            resp = requests.get(
                CKAN_API_INDICADORES,
                params={"id": f"indicadores-criminais-de-{ano}"},
                headers=HEADERS,
                timeout=30,
            )
            if resp.status_code != 200:
                continue
            dados = resp.json()
            if not dados.get("success"):
                continue
            for r in dados["result"].get("resources", []):
                url = r.get("url", "")
                if url.lower().endswith((".xlsx", ".xls")):
                    recursos.append({"ano": ano, "nome": r.get("name", ""), "url": url})
        except (requests.RequestException, ValueError) as e:
            log.warning("Falha ao consultar indicadores criminais de %s: %s", ano, e)
    return recursos


def _baixar_ocorrencias_por_municipio(meses_max: int = 12) -> dict:
    """Baixa os XLSX mensais de indicadores criminais e soma ocorrências por
    município (todas as categorias de crime somadas), últimos `meses_max` meses
    disponíveis."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl não instalado. Rode: pip install openpyxl"
        ) from e

    recursos = _consultar_recursos_indicadores_criminais()
    if not recursos:
        log.warning("Não encontrei arquivos de indicadores criminais no dados.rs.gov.br.")
        return {}

    recursos = sorted(recursos, key=lambda r: (r["ano"], r["nome"]), reverse=True)[:meses_max]

    totais: dict = {}
    for r in recursos:
        try:
            resp = requests.get(r["url"], headers=HEADERS, timeout=60)
            resp.raise_for_status()
            wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
            ws = wb.active
            linhas = list(ws.iter_rows(values_only=True))
        except Exception as e:
            log.warning("Falha ao baixar/ler %s (%s): %s", r["nome"], r["url"], e)
            continue

        idx_header = next(
            (i for i, linha in enumerate(linhas)
             if any(isinstance(c, str) and "MUNIC" in c.upper() for c in linha if c)),
            None,
        )
        if idx_header is None:
            log.warning("Não achei coluna de município em %s", r["nome"])
            continue

        cabecalho = linhas[idx_header]
        col_municipio = next(
            i for i, c in enumerate(cabecalho) if isinstance(c, str) and "MUNIC" in c.upper()
        )

        for linha in linhas[idx_header + 1:]:
            if not linha or linha[col_municipio] is None:
                continue
            municipio = normalizar(str(linha[col_municipio]))
            if municipio not in POPULACAO_MUNICIPIOS_RMPA:
                continue
            soma_linha = sum(
                v for i, v in enumerate(linha)
                if i != col_municipio and isinstance(v, (int, float))
            )
            totais[municipio] = totais.get(municipio, 0) + soma_linha

    return totais


def _classificar_risco(taxa_100k: float, todas_taxas: list) -> str:
    """Classifica em baixo/médio/alto usando tercis da amostra atual (RMPA)."""
    if not todas_taxas:
        return "desconhecido"
    ordenadas = sorted(todas_taxas)
    n = len(ordenadas)
    t1 = ordenadas[max(n // 3 - 1, 0)]
    t2 = ordenadas[max((2 * n) // 3 - 1, 0)]
    if taxa_100k <= t1:
        return "baixo"
    if taxa_100k <= t2:
        return "medio"
    return "alto"


def obter_risco_por_municipio(cache_path: Path) -> dict:
    """Retorna {municipio: {ocorrencias, populacao, taxa_100k, risco}}.

    Usa cache em disco (CACHE_RISCO_DIAS dias) pra não re-baixar os XLSX da
    SSP-RS toda execução -- criminalidade muda devagar, ao contrário dos
    leilões."""
    import json

    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text(encoding="utf-8"))
            idade_dias = (datetime.now() - datetime.fromisoformat(cache["gerado_em"])).days
            if idade_dias < CACHE_RISCO_DIAS:
                log.info("Usando cache de criminalidade (%d dia(s) de idade).", idade_dias)
                return cache["municipios"]
        except Exception as e:
            log.warning("Cache de criminalidade inválido, recalculando: %s", e)

    log.info("Baixando indicadores criminais da SSP-RS (dados.rs.gov.br)...")
    try:
        totais = _baixar_ocorrencias_por_municipio()
    except RuntimeError as e:
        log.warning("%s -- seguindo sem índice de risco.", e)
        return {}

    if not totais:
        log.warning("Não consegui calcular índice de risco -- seguindo sem essa coluna.")
        return {}

    resultado = {}
    for municipio, populacao in POPULACAO_MUNICIPIOS_RMPA.items():
        ocorrencias = totais.get(municipio, 0)
        taxa = (ocorrencias / populacao) * 100_000 if populacao else 0
        resultado[municipio] = {
            "ocorrencias": ocorrencias,
            "populacao": populacao,
            "taxa_100k": round(taxa, 1),
        }

    todas_taxas = [v["taxa_100k"] for v in resultado.values()]
    for municipio in resultado:
        resultado[municipio]["risco"] = _classificar_risco(resultado[municipio]["taxa_100k"], todas_taxas)

    for municipio, info in sorted(resultado.items(), key=lambda kv: -kv[1]["taxa_100k"]):
        log.info(
            "Risco %s: %s ocorrências/12m, %.0f/100k hab. -> %s",
            municipio, info["ocorrencias"], info["taxa_100k"], info["risco"],
        )

    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(
            json.dumps({"gerado_em": datetime.now().isoformat(), "municipios": resultado}, ensure_ascii=False),
            encoding="utf-8",
        )
    except OSError as e:
        log.warning("Não consegui salvar cache de criminalidade: %s", e)

    return resultado

# Coluna que identifica unicamente o imóvel dentro do CSV.
# O nome exato pode variar levemente (ex: "N° do imóvel" vs "Nº do imóvel");
# ajuste aqui se o parsing reclamar na primeira execução real.
ID_COLUMN_CANDIDATES = ["N° do imóvel", "Nº do imóvel", "N do imovel", "N° do Imóvel"]


def baixar_csv(uf: str) -> bytes:
    """Baixa via requests simples (rápido, mas costuma ser bloqueado pelo Radware Bot Manager)."""
    url = BASE_URL.format(uf=uf.upper())
    log.info("Baixando lista de %s: %s", uf, url)
    resp = requests.get(url, headers=HEADERS, timeout=60)
    log.info(
        "Resposta: status=%s content-type=%s content-length=%s bytes-recebidos=%d",
        resp.status_code,
        resp.headers.get("Content-Type"),
        resp.headers.get("Content-Length"),
        len(resp.content),
    )
    resp.raise_for_status()
    return resp.content


def parece_desafio_radware(conteudo: bytes) -> bool:
    amostra = conteudo[:2000].lower()
    return b"radware" in amostra or b"bot manager" in amostra or b"captcha" in amostra


def baixar_csv_via_navegador(uf: str, headless: bool = True, tentativas: int = 3) -> bytes:
    """Baixa o CSV usando um Chrome de verdade (via Playwright).

    Necessário porque o site usa Radware Bot Manager, que bloqueia requisições
    HTTP "cruas" (sem JS, sem fingerprint de navegador). Um Chrome de verdade
    resolve o desafio automaticamente ao carregar a página.

    O arquivo CSV é servido com cabeçalho de download (Content-Disposition),
    então o Chrome trata a navegação como um download em vez de renderizar a
    página -- por isso primeiro "aquecemos" a sessão numa página normal do
    site (pra passar o desafio do Radware e ganhar os cookies), e só depois
    disparamos o download do CSV em si, capturando o evento de download.

    Requer: pip install playwright && playwright install chromium
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as e:
        raise RuntimeError(
            "Playwright não está instalado. Rode:\n"
            "  pip install playwright\n"
            "  playwright install chromium"
        ) from e

    url_csv = BASE_URL.format(uf=uf.upper())
    url_aquecimento = "https://venda-imoveis.caixa.gov.br/sistema/busca-imovel.asp"
    log.info("Baixando lista de %s via navegador: %s", uf, url_csv)

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled"],
        )
        contexto = browser.new_context(
            user_agent=HEADERS["User-Agent"],
            locale="pt-BR",
            accept_downloads=True,
        )
        pagina = contexto.new_page()

        # 1) Aquece a sessão numa página normal (resolve o desafio do Radware, ganha cookies)
        log.info("Aquecendo sessão em %s ...", url_aquecimento)
        pagina.goto(url_aquecimento, wait_until="networkidle", timeout=30000)
        pagina.wait_for_timeout(3000)

        conteudo = b""
        for tentativa in range(1, tentativas + 1):
            try:
                # 2) Dispara o download do CSV em si (o goto vai "falhar" com
                #    "Download is starting" -- isso é esperado, o download
                #    real é capturado pelo expect_download)
                with pagina.expect_download(timeout=30000) as download_info:
                    try:
                        pagina.goto(url_csv, timeout=15000)
                    except Exception:
                        pass  # esperado: goto interrompe por causa do download
                download = download_info.value
                caminho_tmp = download.path()
                conteudo = Path(caminho_tmp).read_bytes()
            except Exception as e:
                log.warning("Tentativa %d: não veio como download (%s), tentando ler como página...", tentativa, e)
                resposta = pagina.goto(url_csv, wait_until="networkidle", timeout=30000)
                conteudo = resposta.body() if resposta else b""

            if conteudo and not parece_desafio_radware(conteudo):
                log.info("Desafio passado na tentativa %d (bytes: %d)", tentativa, len(conteudo))
                break
            log.warning("Tentativa %d: ainda recebendo o desafio do Radware, aguardando e tentando de novo...", tentativa)
            pagina.wait_for_timeout(4000)
        else:
            log.warning("Não consegui passar do desafio do Radware após %d tentativas.", tentativas)

        browser.close()

    return conteudo


def parsear_csv(conteudo: bytes) -> list[dict]:
    """Faz o parsing do CSV da Caixa.

    O arquivo real vem assim (confirmado em 2026-07):
        Linha 1: "Lista de Imóveis da Caixa;;Data de geração:;DD/MM/AAAA;;;;;;;"  <- lixo, ignorar
        Linha 2: " Nº do imóvel;UF;Cidade;Bairro;Endereço;Preço;..."              <- cabeçalho real
        Linha 3: em branco
        Linha 4+: dados, separados por ';', campos com espaços sobrando

    Por isso não dá pra usar csv.Sniffer nem assumir que a primeira linha é o
    cabeçalho -- é preciso achar a linha que contém "Cidade" e "UF".
    """
    texto = conteudo.decode("latin-1", errors="replace")
    linhas_texto = texto.splitlines()

    idx_header = None
    for i, linha in enumerate(linhas_texto):
        if "Cidade" in linha and "UF" in linha:
            idx_header = i
            break

    if idx_header is None:
        log.warning("Não encontrei a linha de cabeçalho esperada (com 'Cidade' e 'UF'). "
                     "O formato do arquivo pode ter mudado. Tentando a partir da linha 0.")
        log.warning("Amostra do conteúdo bruto recebido (primeiros 1000 chars): %r", texto[:1000])
        idx_header = 0

    texto_util = "\n".join(l for l in linhas_texto[idx_header:] if l.strip())
    leitor = csv.DictReader(io.StringIO(texto_util), delimiter=";")

    linhas = []
    for row in leitor:
        # limpa espaços em branco de chaves e valores (o CSV da Caixa vem cheio deles)
        row_limpo = {(k or "").strip(): (v or "").strip() for k, v in row.items() if k}
        if any(row_limpo.values()):
            linhas.append(row_limpo)

    log.info("Parseadas %d linhas", len(linhas))
    return linhas


def filtrar_por_cidade(linhas: list[dict], cidades: list[str]) -> list[dict]:
    """Mantém só as linhas cuja cidade bate com alguma da lista.

    Usa a coluna 'Cidade' quando existe (caso normal do CSV da Caixa). Se por
    algum motivo essa coluna não existir, cai pro modo antigo (procura a
    cidade em qualquer campo da linha).
    """
    cidades_norm = [normalizar(c) for c in cidades]
    tem_coluna_cidade = linhas and "Cidade" in linhas[0]

    filtradas = []
    for row in linhas:
        if tem_coluna_cidade:
            alvo = normalizar(row.get("Cidade", ""))
            bate = any(alvo == cidade or cidade in alvo for cidade in cidades_norm)
        else:
            texto_linha = normalizar(" | ".join(str(v) for v in row.values() if v))
            bate = any(cidade in texto_linha for cidade in cidades_norm)
        if bate:
            filtradas.append(row)

    log.info("Filtro de cidade: %d de %d linhas mantidas", len(filtradas), len(linhas))
    return filtradas


def achar_coluna_id(linhas: list[dict]) -> str:
    if not linhas:
        return ""
    colunas = linhas[0].keys()
    for candidata in ID_COLUMN_CANDIDATES:
        if candidata in colunas:
            return candidata
    # fallback: primeira coluna
    primeira = next(iter(colunas))
    log.warning("Coluna de ID não encontrada nos candidatos, usando fallback: %r", primeira)
    return primeira


def salvar_snapshot(linhas: list[dict], uf: str, outdir: Path) -> Path:
    outdir.mkdir(parents=True, exist_ok=True)
    data_str = datetime.now().strftime("%Y-%m-%d")
    caminho = outdir / f"{uf.upper()}_{data_str}.csv"
    if linhas:
        with caminho.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(linhas[0].keys()))
            writer.writeheader()
            writer.writerows(linhas)
    log.info("Snapshot salvo em %s", caminho)
    return caminho


def carregar_snapshot_anterior(uf: str, outdir: Path, data_atual: str) -> list[dict] | None:
    arquivos = sorted(outdir.glob(f"{uf.upper()}_*.csv"))
    anteriores = [a for a in arquivos if a.stem != f"{uf.upper()}_{data_atual}"]
    if not anteriores:
        return None
    ultimo = anteriores[-1]
    log.info("Comparando com snapshot anterior: %s", ultimo)
    with ultimo.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def comparar(atuais: list[dict], anteriores: list[dict] | None, id_col: str):
    if anteriores is None:
        return atuais, []  # tudo é "novo" na primeira execução
    ids_atuais = {row.get(id_col) for row in atuais}
    ids_anteriores = {row.get(id_col) for row in anteriores}
    novos_ids = ids_atuais - ids_anteriores
    removidos_ids = ids_anteriores - ids_atuais
    novos = [row for row in atuais if row.get(id_col) in novos_ids]
    removidos = [row for row in anteriores if row.get(id_col) in removidos_ids]
    return novos, removidos


def _parar_valor_brl(texto: str) -> float:
    """Converte '147.398,67' (formato BR: ponto=milhar, vírgula=decimal) em float."""
    if not texto:
        return 0.0
    limpo = texto.strip().replace(".", "").replace(",", ".")
    try:
        return float(limpo)
    except ValueError:
        return 0.0


def _parar_desconto(texto: str) -> float:
    """Desconto já vem como '48.10' (ponto decimal padrão)."""
    try:
        return float((texto or "0").strip())
    except ValueError:
        return 0.0


def extrair_tipo(descricao: str) -> str:
    d = descricao.lower()
    if 'apartamento' in d or 'apto' in d: return 'Apartamento'
    if 'casa' in d or 'sobrado' in d: return 'Casa'
    if 'comercial' in d or 'sala' in d or 'galpão' in d or 'galpao' in d or 'loja' in d or 'prédio' in d or 'predio' in d or 'pavilhão' in d: return 'Comercial'
    if 'rural' in d or 'fazenda' in d or 'sítio' in d or 'sitio' in d or 'chácara' in d: return 'Rural'
    if 'terreno' in d or 'lote' in d or 'gleba' in d: return 'Terreno'
    if 'box' in d or 'vaga' in d or 'garagem' in d: return 'Box/Garagem'
    return 'Outro'


def _montar_registro_js(row: dict, ids_novos: set[str], id_col: str, risco_por_municipio: dict, risco_bairro_poa: dict | None = None) -> dict:
    preco = _parar_valor_brl(row.get("Preço", ""))
    avaliacao = _parar_valor_brl(row.get("Valor de avaliação", ""))
    desconto = _parar_desconto(row.get("Desconto", ""))
    cidade = row.get("Cidade", "")
    bairro_norm = normalizar(row.get("Bairro", ""))
    cidade_norm = normalizar(cidade)
    info_risco = None
    if risco_bairro_poa:
        bairros_cidade = risco_bairro_poa.get(cidade_norm, {})
        if bairros_cidade:
            info_risco = bairros_cidade.get(bairro_norm)
    if info_risco is None:
        info_risco = risco_por_municipio.get(cidade_norm, {})
    descricao = row.get("Descrição", "")
    import re
    area_match = re.search(r'([\d\.]+)\s*de área privativa', descricao)
    area_privativa = float(area_match.group(1)) if area_match else 0.0
    preco_m2 = preco / area_privativa if area_privativa > 0 else 0.0

    return {
        "id": row.get(id_col, ""),
        "cidade": cidade_norm.title(),
        "bairro": bairro_norm.title(),
        "tipo": extrair_tipo(descricao),
        "endereco": row.get("Endereço", ""),
        "preco": preco,
        "avaliacao": avaliacao,
        "desconto": desconto,
        "financiamento": row.get("Financiamento", ""),
        "modalidade": row.get("Modalidade de venda", ""),
        "descricao": descricao,
        "link": row.get("Link de acesso", ""),
        "origem": row.get("Origem", "Caixa"),
        "novo": row.get(id_col, "") in ids_novos,
        "risco": info_risco.get("risco", "desconhecido"),
        "taxa_criminalidade": info_risco.get("taxa_100k"),
        "area_privativa": area_privativa,
        "preco_m2": preco_m2,
    }


def gerar_html(
    uf: str,
    data_atual: str,
    novos: list[dict],
    removidos: list[dict],
    todas: list[dict],
    cidades: list[str] | None,
    caminho_saida: Path,
    risco_por_municipio: dict | None = None,
    risco_bairro_poa: dict | None = None,
    historico_precos: dict | None = None,
):
    import json
    import re

    risco_por_municipio = risco_por_municipio or {}
    historico_precos = historico_precos or {}
    id_col = achar_coluna_id(todas) if todas else ""
    ids_novos = {row.get(id_col) for row in novos} if todas else set()
    
    ruas_contagem = {}
    rua_por_imovel = {}
    for row in todas:
        imovel_id = row.get(id_col, "")
        endereco = row.get("Endereço", "")
        cidade = row.get("Cidade", "").title()
        rua = re.split(r',| - | N\.| n\.| S/N| s/n| Nº| nº', endereco)[0].strip().upper()
        if len(rua) < 3: rua = endereco.strip().upper()
        chave = f"{rua} | {cidade}"
        rua_por_imovel[imovel_id] = rua
        ruas_contagem[chave] = ruas_contagem.get(chave, 0) + 1
        
    registros = []
    for row in todas:
        imovel_id = row.get(id_col, "")
        reg = _montar_registro_js(row, ids_novos, id_col, risco_por_municipio, risco_bairro_poa)
        
        hist = historico_precos.get(imovel_id)
        if hist and len(hist.get("precos", [])) > 1:
            primeiro_preco = hist["precos"][0]
            ultimo_preco = hist["precos"][-1]
            if ultimo_preco < primeiro_preco - 0.01:
                reg["queda_preco"] = primeiro_preco - ultimo_preco
                
        if hist and len(hist.get("datas", [])) > 0:
            from datetime import datetime
            try:
                d_pri = datetime.strptime(hist["datas"][0], "%Y-%m-%d")
                d_hoje = datetime.strptime(data_atual, "%Y-%m-%d")
                reg["idade_dias"] = (d_hoje - d_pri).days
            except:
                reg["idade_dias"] = 0
        else:
            reg["idade_dias"] = 0
                
        cidade = row.get("Cidade", "").title()
        rua = rua_por_imovel.get(imovel_id, "")
        chave = f"{rua} | {cidade}"
        cont = ruas_contagem.get(chave, 1)
        if cont > 1:
            reg["contagem_rua"] = cont
            reg["nome_rua"] = rua
            
        registros.append(reg)

    registros.sort(key=lambda r: r["desconto"], reverse=True)

    bairros = sorted({r["bairro"] for r in registros if r["bairro"]})
    modalidades = sorted({r["modalidade"] for r in registros if r["modalidade"]})
    cidades_disp = sorted({r["cidade"] for r in registros if r["cidade"]})
    tipos = sorted({r["tipo"] for r in registros if r.get("tipo")})
    tem_dados_risco = any(r["risco"] != "desconhecido" for r in registros)

    desconto_medio = sum(r["desconto"] for r in registros) / len(registros) if registros else 0
    melhor = max(registros, key=lambda r: r["desconto"]) if registros else None

    dados_json = json.dumps(registros, ensure_ascii=False).replace("</script>", "<\\/script>")
    filtro_txt = ", ".join(cidades) if cidades else "sem filtro (todas as cidades)"

    opcoes_bairro = "".join(f'<option value="{b}">{b}</option>' for b in bairros)
    opcoes_modalidade = "".join(f'<option value="{m}">{m}</option>' for m in modalidades)
    opcoes_cidade = "".join(f'<option value="{c}">{c}</option>' for c in cidades_disp)
    opcoes_tipo = "".join(f'<option value="{t}">{t}</option>' for t in tipos)

    removidos_html = ""
    if removidos:
        itens = "".join(
            f"<li>{r.get('Bairro','').title()} — {r.get('Endereço','')}</li>" for r in removidos
        )
        removidos_html = f"""
        <details class="removidos">
          <summary>{len(removidos)} imóvel(is) saíram da lista desde ontem</summary>
          <ul>{itens}</ul>
        </details>"""

    resumo_melhor = ""
    if melhor:
        resumo_melhor = (
            f'{melhor["desconto"]:.0f}% em {melhor["bairro"] or melhor["cidade"]}'
        )

    filtro_risco_html = ""
    aviso_risco_html = ""
    if tem_dados_risco:
        filtro_risco_html = """
    <label>Risco
      <select id="filtro-risco"><option value="">Todos</option><option value="baixo">Baixo</option><option value="medio">Médio</option><option value="alto">Alto</option></select>
    </label>"""
        if risco_bairro_poa:
            aviso_risco_html = (
                '<div class="aviso-risco">⚠ Risco calculado por <b>bairro</b> em Porto Alegre '
                '(ocorrências/100 mil hab., dados SSP-RS + população Censo IBGE 2022) e por '
                '<b>cidade</b> nas demais localidades da região metropolitana. É um proxy de '
                'segurança pública, não um índice de retorno do investimento.</div>'
            )
        else:
            aviso_risco_html = (
                '<div class="aviso-risco">⚠ O risco é calculado por <b>cidade</b> (dados SSP-RS, '
                'ocorrências/100 mil hab.), não por bairro — a fonte pública não disponibilizou essa '
                'granularidade nesta execução. Dentro de Porto Alegre, todos os bairros recebem '
                'o mesmo nível. Trate como triagem inicial, não como avaliação do endereço específico.</div>'
            )

    html = f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Leilões Caixa · {uf} · {data_atual}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@500;600;700&family=IBM+Plex+Mono:wght@400;500;600&family=Inter:wght@400;500;600&display=swap" rel="stylesheet">
<style>
  :root {{
    --bg: #0a0b0d;
    --panel: #131519;
    --panel-2: #191c21;
    --border: #262a31;
    --text: #eceef1;
    --muted: #8790a0;
    --gold: #d4a35a;
    --gold-dim: #8a713f;
    --good: #4ea87a;
    --mid: #c9a227;
    --low: #6b7280;
    --novo-bg: #1a2a20;
    --novo-border: #2f6b45;
  }}
  * {{ box-sizing: border-box; }}
  body {{
    margin: 0; background: var(--bg); color: var(--text);
    font-family: 'Inter', -apple-system, sans-serif;
    -webkit-font-smoothing: antialiased;
  }}
  .num {{ font-family: 'IBM Plex Mono', monospace; }}
  .container {{ max-width: 1320px; margin: 0 auto; padding: 28px 20px 60px; }}

  header.top {{ display: flex; justify-content: space-between; align-items: flex-end; flex-wrap: wrap; gap: 16px; margin-bottom: 22px; }}
  h1 {{
    font-family: 'Barlow Condensed', sans-serif; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.04em; font-size: 34px; margin: 0; color: var(--text);
  }}
  h1 span {{ color: var(--gold); }}
  .subtitulo {{ color: var(--muted); font-size: 13px; margin-top: 4px; }}

  .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 10px; margin-bottom: 22px; }}
  .stat {{ background: var(--panel); border: 1px solid var(--border); border-radius: 8px; padding: 14px 16px; }}
  .stat .valor {{ font-family: 'Barlow Condensed', sans-serif; font-size: 26px; font-weight: 600; line-height: 1; }}
  .stat .valor.gold {{ color: var(--gold); }}
  .stat .rotulo {{ color: var(--muted); font-size: 11px; text-transform: uppercase; letter-spacing: 0.05em; margin-top: 5px; }}

  .toolbar {{
    display: flex; flex-wrap: wrap; gap: 8px; align-items: center;
    background: var(--panel); border: 1px solid var(--border); border-radius: 8px;
    padding: 12px; margin-bottom: 16px; position: sticky; top: 12px; z-index: 5;
  }}
  .toolbar input[type="text"], .toolbar input[type="number"], .toolbar select {{
    background: var(--panel-2); border: 1px solid var(--border); color: var(--text);
    border-radius: 6px; padding: 8px 10px; font-size: 13px; font-family: inherit;
  }}
  .busca-wrapper {{ flex: 1 1 220px; min-width: 160px; position: relative; display: flex; align-items: center; }}
  .toolbar input[type="text"] {{ flex: 1; width: 100%; padding-right: 30px; }}
  #limpar-busca {{
    position: absolute; right: 8px; background: transparent; border: none;
    color: var(--muted); cursor: pointer; font-size: 18px; font-weight: bold;
    display: none; padding: 0 4px; line-height: 1;
  }}
  #limpar-busca:hover {{ color: var(--text); }}
  .toolbar input[type="number"] {{ width: 110px; }}
  .toolbar select {{ width: 140px; text-overflow: ellipsis; overflow: hidden; white-space: nowrap; }}
  .toolbar label {{ font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing: 0.04em; display: flex; flex-direction: column; gap: 4px; }}
  .toolbar .desconto-min {{ display: flex; align-items: center; gap: 8px; }}
  #desconto-min-valor {{ display: inline-block; width: 32px; text-align: right; }}
  .toolbar input[type="range"] {{ accent-color: var(--gold); width: 110px; }}
  .contagem {{ margin-left: auto; color: var(--muted); font-size: 12px; }}
  .contagem b {{ color: var(--text); }}

  .tabela-wrap {{ overflow-x: auto; border: 1px solid var(--border); border-radius: 8px; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13px; min-width: 900px; }}
  th, td {{ padding: 10px 12px; text-align: left; border-bottom: 1px solid var(--border); white-space: nowrap; }}
  th {{
    background: var(--panel-2); color: var(--muted); font-weight: 600; font-size: 11px;
    text-transform: uppercase; letter-spacing: 0.04em; position: sticky; top: 0; cursor: pointer;
    user-select: none;
  }}
  th:hover {{ color: var(--gold); }}
  th.ativo {{ color: var(--gold); }}
  tbody tr {{ background: var(--panel); }}
  tbody tr:hover {{ background: var(--panel-2); }}
  tbody tr.novo {{ background: var(--novo-bg); }}
  tbody tr.novo:hover {{ background: #223829; }}
  td.endereco {{ white-space: normal; max-width: 260px; }}
  td.endereco .bairro {{ font-weight: 600; }}
  td.endereco .rua {{ color: var(--muted); font-size: 12px; display: block; text-decoration: none; }}
  td.endereco a.rua:hover {{ text-decoration: underline; color: var(--gold); }}
  .badge-novo {{
    background: var(--good); color: #06170e; font-size: 10px; font-weight: 700;
    padding: 2px 6px; border-radius: 4px; text-transform: uppercase; letter-spacing: 0.03em; margin-left: 6px;
  }}
  .badge-rua {{
    display: inline-block; margin-top: 4px; padding: 2px 6px;
    background: #2a3a5a; color: #a4c2f4; font-size: 10px; font-weight: 600;
    border-radius: 4px; cursor: pointer; border: 1px solid #3b507a;
  }}
  .badge-rua:hover {{ background: #3b507a; }}
  .badge-idade {{ display: inline-block; padding: 2px 6px; border-radius: 4px; font-size: 10px; font-weight: 700; margin-left: 6px; vertical-align: middle; }}
  .idade-novo {{ background: var(--good); color: #06170e; }}
  .idade-encalhado {{ background: #5c1818; color: #f4a4a4; }}
  .idade-morno {{ background: var(--panel-2); color: var(--muted); }}
  .queda-preco {{
    color: var(--good); font-size: 11px; font-weight: 600; margin-top: 2px;
  }}
  .desconto-cel {{ display: flex; flex-direction: column; gap: 4px; min-width: 90px; }}
  .desconto-num {{ font-weight: 600; }}
  .barra {{ height: 4px; border-radius: 2px; background: var(--border); overflow: hidden; }}
  .barra > i {{ display: block; height: 100%; border-radius: 2px; }}
  .modalidade-tag {{
    font-size: 11px; padding: 3px 8px; border-radius: 100px; border: 1px solid var(--border);
    color: var(--muted); white-space: nowrap;
  }}
  .risco-tag {{
    font-size: 11px; padding: 3px 8px; border-radius: 100px; font-weight: 600; white-space: nowrap;
  }}
  .risco-baixo {{ background: rgba(78,168,122,0.15); color: var(--good); }}
  .risco-medio {{ background: rgba(201,162,39,0.15); color: var(--mid); }}
  .risco-alto {{ background: rgba(200,90,80,0.15); color: #d97a6c; }}
  .risco-desconhecido {{ background: transparent; color: var(--muted); border: 1px dashed var(--border); }}
  .aviso-risco {{
    font-size: 12px; color: var(--muted); background: var(--panel); border: 1px solid var(--border);
    border-radius: 8px; padding: 10px 14px; margin-bottom: 14px; line-height: 1.5;
  }}
  a.ver-btn {{
    display: inline-block; color: var(--bg); background: var(--gold); font-weight: 600;
    font-size: 12px; padding: 5px 10px; border-radius: 6px; text-decoration: none; white-space: nowrap;
  }}
  a.ver-btn:hover {{ background: #e3b671; }}

  .vazio {{ color: var(--muted); font-style: italic; padding: 24px; text-align: center; }}

  .removidos {{ margin-top: 20px; color: var(--muted); font-size: 13px; }}
  .removidos summary {{ cursor: pointer; color: var(--muted); }}
  .removidos ul {{ margin: 8px 0 0; padding-left: 18px; }}

  footer {{ margin-top: 30px; color: var(--muted); font-size: 11px; text-align: center; }}

  @media (max-width: 640px) {{
    h1 {{ font-size: 26px; }}
    .toolbar {{ position: static; }}
  }}
</style>
</head>
<body>
<div class="container">
  <header class="top">
    <div>
      <h1>Leilões <span>Caixa</span> — {uf}</h1>
      <div class="subtitulo">Atualizado em {data_atual} · Região: {filtro_txt}</div>
    </div>
  </header>

  <div class="stats">
    <div class="stat"><div class="valor">{len(todas)}</div><div class="rotulo">Ativos hoje</div></div>
    <div class="stat"><div class="valor gold">{len(novos)}</div><div class="rotulo">Novos hoje</div></div>
    <div class="stat"><div class="valor">{desconto_medio:.0f}%</div><div class="rotulo">Desconto médio</div></div>
    <div class="stat"><div class="valor gold">{resumo_melhor or '—'}</div><div class="rotulo">Maior desconto</div></div>
  </div>

  <div class="toolbar">
    <div class="busca-wrapper">
      <input type="text" id="busca" placeholder="Buscar por endereço, bairro...">
      <button id="limpar-busca" title="Limpar busca">&times;</button>
    </div>
    <label>Cidade
      <select id="filtro-cidade"><option value="">Todas</option>{opcoes_cidade}</select>
    </label>
    <label>Bairro
      <select id="filtro-bairro"><option value="">Todos</option>{opcoes_bairro}</select>
    </label>
    <label>Tipo
      <select id="filtro-tipo"><option value="">Todos</option>{opcoes_tipo}</select>
    </label>
    <label>Modalidade
      <select id="filtro-modalidade"><option value="">Todas</option>{opcoes_modalidade}</select>
    </label>
    <label>Só novos hoje
      <select id="filtro-novos"><option value="">Todos</option><option value="1">Sim</option></select>
    </label>{filtro_risco_html}
    <label>Preço Mín. (R$)
      <input type="number" id="preco-min" min="0" step="10000" placeholder="0">
    </label>
    <label>Preço Máx. (R$)
      <input type="number" id="preco-max" min="0" step="10000" placeholder="Máx">
    </label>
    <div class="desconto-min">
      <label>Desconto mín. <span id="desconto-min-valor">0%</span>
        <input type="range" id="desconto-min" min="0" max="90" value="0" step="5">
      </label>
    </div>
    <span class="contagem"><b id="contagem-num">{len(todas)}</b> imóveis</span>
  </div>

  {aviso_risco_html}

  <div class="tabela-wrap">
    <table id="tabela">
      <thead>
        <tr>
          <th data-col="bairro">Local</th>
          <th data-col="preco" class="num">Preço</th>
          <th data-col="avaliacao" class="num">Avaliação</th>
          <th data-col="preco_m2" class="num">R$/m²</th>
          <th data-col="desconto" class="ativo">Desconto</th>
          <th data-col="risco">Risco</th>
          <th data-col="modalidade">Modalidade</th>
          <th></th>
        </tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
    <div id="vazio" class="vazio" style="display:none">Nenhum imóvel bate com esse filtro.</div>
  </div>

  {removidos_html}

  <footer>Fonte: venda-imoveis.caixa.gov.br · dados públicos da Caixa Econômica Federal</footer>
</div>

<script>
const DADOS = {dados_json};

const fmtBRL = (v) => v.toLocaleString('pt-BR', {{style:'currency', currency:'BRL', maximumFractionDigits:0}});

function corDesconto(d) {{
  if (d >= 50) return 'var(--good)';
  if (d >= 30) return 'var(--mid)';
  return 'var(--low)';
}}

let ordenarPor = 'desconto';
let ordemAsc = false;

function aplicarFiltros() {{
  const inputBuscaEl = document.getElementById('busca');
  const busca = inputBuscaEl.value.trim().toLowerCase();
  const btnLimpar = document.getElementById('limpar-busca');
  if (btnLimpar) {{
    btnLimpar.style.display = inputBuscaEl.value ? 'block' : 'none';
  }}
  
  const f = {{
    busca: busca,
    cidade: document.getElementById('filtro-cidade').value,
    bairro: document.getElementById('filtro-bairro').value,
    tipo: document.getElementById('filtro-tipo') ? document.getElementById('filtro-tipo').value : '',
    modalidade: document.getElementById('filtro-modalidade').value,
    soNovos: document.getElementById('filtro-novos').value,
    risco: document.getElementById('filtro-risco') ? document.getElementById('filtro-risco').value : '',
    descMin: parseInt(document.getElementById('desconto-min').value, 10) || 0,
    precoMin: parseFloat(document.getElementById('preco-min').value) || 0,
    precoMax: document.getElementById('preco-max').value ? parseFloat(document.getElementById('preco-max').value) : Infinity
  }};

  const isMatch = (r, exclude) => {{
    if (exclude !== 'busca' && f.busca && !(r.endereco.toLowerCase().includes(f.busca) || r.bairro.toLowerCase().includes(f.busca))) return false;
    if (exclude !== 'cidade' && f.cidade && r.cidade !== f.cidade) return false;
    if (exclude !== 'bairro' && f.bairro && r.bairro !== f.bairro) return false;
    if (exclude !== 'tipo' && f.tipo && r.tipo !== f.tipo) return false;
    if (exclude !== 'modalidade' && f.modalidade && r.modalidade !== f.modalidade) return false;
    if (exclude !== 'soNovos' && f.soNovos === '1' && !r.novo) return false;
    if (exclude !== 'risco' && f.risco && r.risco !== f.risco) return false;
    if (r.desconto < f.descMin) return false;
    if (r.preco < f.precoMin || r.preco > f.precoMax) return false;
    return true;
  }};

  const rebuildSelect = (id, prop) => {{
    const el = document.getElementById(id);
    if (!el) return;
    const currentVal = el.value;
    const options = new Set();
    DADOS.forEach(r => {{
      if (isMatch(r, prop) && r[prop]) options.add(r[prop]);
    }});
    const sorted = [...options].sort();
    let html = '<option value="">' + (id === 'filtro-cidade' ? 'Todas' : 'Todos') + '</option>';
    sorted.forEach(val => {{
      html += `<option value="${{val}}">${{val}}</option>`;
    }});
    el.innerHTML = html;
    if (sorted.includes(currentVal)) {{
      el.value = currentVal;
    }} else {{
      el.value = '';
      f[prop] = '';
    }}
  }};

  rebuildSelect('filtro-cidade', 'cidade');
  rebuildSelect('filtro-bairro', 'bairro');
  rebuildSelect('filtro-tipo', 'tipo');
  rebuildSelect('filtro-modalidade', 'modalidade');
  rebuildSelect('filtro-risco', 'risco');

  let filtrados = DADOS.filter(r => isMatch(r, null));

  filtrados.sort((a, b) => {{
    let va = a[ordenarPor], vb = b[ordenarPor];
    if (typeof va === 'string') {{ va = va.toLowerCase(); vb = vb.toLowerCase(); }}
    if (va < vb) return ordemAsc ? -1 : 1;
    if (va > vb) return ordemAsc ? 1 : -1;
    return 0;
  }});

  renderizar(filtrados);
}}

function rotuloRisco(r) {{
  const rotulos = {{baixo: 'Baixo', medio: 'Médio', alto: 'Alto', desconhecido: '—'}};
  return rotulos[r] || '—';
}}

function renderizar(lista) {{
  const tbody = document.getElementById('tbody');
  const vazio = document.getElementById('vazio');
  document.getElementById('contagem-num').textContent = lista.length;

  if (lista.length === 0) {{
    tbody.innerHTML = '';
    vazio.style.display = 'block';
    return;
  }}
  vazio.style.display = 'none';

  tbody.innerHTML = lista.map(r => `
    <tr class="${{r.novo ? 'novo' : ''}}">
      <td class="endereco">
        <span class="bairro">
          <span class="badge-origem" style="background: ${{r.origem === 'Caixa' ? '#005ca9' : (r.origem === 'Mega Leilões' ? '#b22222' : '#2e8b57')}}; color: white; padding: 2px 6px; border-radius: 4px; font-size: 10px; font-weight: bold; margin-right: 6px;">${{r.origem}}</span>
          ${{r.bairro || r.cidade}}
          ${{r.idade_dias !== undefined && r.idade_dias <= 3 ? '<span class="badge-idade idade-novo">🔥 NOVO</span>' : (r.idade_dias >= 120 ? `<span class="badge-idade idade-encalhado">⏳ ${{r.idade_dias}} DIAS</span>` : `<span class="badge-idade idade-morno">⏳ ${{r.idade_dias}} DIAS</span>`)}}
        </span>
        <a href="https://www.google.com/maps/search/?api=1&query=${{encodeURIComponent(r.endereco + ', ' + r.cidade)}}" target="_blank" class="rua" title="Ver no Google Maps">${{r.endereco}}</a>
        ${{r.contagem_rua ? `<span class="badge-rua" onclick="document.getElementById('busca').value='${{r.nome_rua}}'; aplicarFiltros()" title="Filtrar por esta rua">📍 ${{r.contagem_rua}} na mesma rua</span>` : ''}}
      </td>
      <td class="num">
        ${{fmtBRL(r.preco)}}
        ${{r.queda_preco ? `<div class="queda-preco">📉 Caiu ${{fmtBRL(r.queda_preco)}}</div>` : ''}}
      </td>
      <td class="num">${{fmtBRL(r.avaliacao)}}</td>
      <td class="num">${{r.preco_m2 > 0 ? fmtBRL(r.preco_m2) + '/m²' : '—'}}</td>
      <td>
        <div class="desconto-cel">
          <span class="desconto-num num" style="color:${{corDesconto(r.desconto)}}">${{r.desconto.toFixed(0)}}%</span>
          <div class="barra"><i style="width:${{Math.min(r.desconto,100)}}%; background:${{corDesconto(r.desconto)}}"></i></div>
        </div>
      </td>
      <td><span class="risco-tag risco-${{r.risco}}" title="${{r.taxa_criminalidade != null ? r.taxa_criminalidade + ' ocorrências/100k hab. (' + r.cidade + ')' : 'Sem dado'}}">${{rotuloRisco(r.risco)}}</span></td>
      <td><span class="modalidade-tag">${{r.modalidade}}</span></td>
      <td><a class="ver-btn" href="${{r.link}}" target="_blank" rel="noopener">Ver no site →</a></td>
    </tr>
  `).join('');
}}

document.querySelectorAll('th[data-col]').forEach(th => {{
  th.addEventListener('click', () => {{
    const col = th.dataset.col;
    if (ordenarPor === col) {{ ordemAsc = !ordemAsc; }}
    else {{ ordenarPor = col; ordemAsc = false; }}
    document.querySelectorAll('th[data-col]').forEach(t => t.classList.remove('ativo'));
    th.classList.add('ativo');
    aplicarFiltros();
  }});
}});

document.getElementById('desconto-min').addEventListener('input', (e) => {{
  document.getElementById('desconto-min-valor').textContent = e.target.value + '%';
  aplicarFiltros();
}});
['busca','filtro-cidade','filtro-bairro','filtro-tipo','filtro-modalidade','filtro-novos','filtro-risco','preco-min','preco-max'].forEach(id => {{
  const el = document.getElementById(id);
  if (el) el.addEventListener('input', aplicarFiltros);
}});

const btnLimpar = document.getElementById('limpar-busca');
if (btnLimpar) {{
  btnLimpar.addEventListener('click', () => {{
    document.getElementById('busca').value = '';
    aplicarFiltros();
  }});
}}

aplicarFiltros();
</script>
</body>
</html>"""
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    caminho_saida.write_text(html, encoding="utf-8")
    log.info("Relatório HTML salvo em %s", caminho_saida)


def _descobrir_zips_ocorrencias_desagregadas(headless: bool = True, max_arquivos: int = 2) -> list:
    """Reusa o navegador pra achar os links dos ZIPs de ocorrências desagregadas
    (artigo 3º, Lei 15.610/2021 -- tem campo de bairro) na página de dados
    abertos da SSP-RS. Pega os mais recentes (ano atual + ano anterior)."""
    links = explorar_dados_abertos_ssp(headless=headless)
    candidatos = [
        l for l in links
        if "ocorrencias" in l["href"].lower() or "ocorrencias" in l["texto"].lower()
    ]
    # heurística: nome/texto do link costuma trazer o ano (ex: "2026 (janeiro a maio)")
    candidatos_ordenados = sorted(candidatos, key=lambda l: l["texto"], reverse=True)
    return candidatos_ordenados[:max_arquivos]


def _extrair_csv_de_zip(conteudo_zip: bytes):
    import zipfile

    with zipfile.ZipFile(io.BytesIO(conteudo_zip)) as z:
        nomes_csv = [n for n in z.namelist() if n.lower().endswith((".csv", ".txt"))]
        if not nomes_csv:
            return None
        with z.open(nomes_csv[0]) as f:
            return f.read()


# Variações de nome de bairro como aparecem nos BOs da SSP-RS vs. nome oficial do IBGE/Prefeitura.
# Alimentado pelas maiores discrepâncias encontradas ao processar os microdados.
_ALIASES_BAIRRO_POA = {
    "PASSO D'AREIA": "PASSO DA AREIA",
    "PASSO D AREIA": "PASSO DA AREIA",
    "LOMBA PINHEIRO": "LOMBA DO PINHEIRO",
    "ABERTA MORROS": "ABERTA DOS MORROS",
    "JARDIM DONA LEOPOLDINA": "JARDIM LEOPOLDINA",
    "MONT SERRAT": "MONT'SERRAT",
    "MONTSERRAT": "MONT'SERRAT",
    "CORONEL APARICIO BORGES": "CORONEL APARÍCIO BORGES",
    "CHAPEU DO SOL": "CHAPÉU DO SOL",
    "TRES FIGUEIRAS": "TRÊS FIGUEIRAS",
    "SAO SEBASTIAO": "SÃO SEBASTIÃO",
    "SAO GERALDO": "SÃO GERALDO",
    "SAO JOAO": "SÃO JOÃO",
    "SAO CAETANO": "SÃO CAETANO",
}


def _resolver_bairro_oficial(bruto_norm: str, oficiais_norm: set, oficiais_lista: list) -> str | None:
    """Tenta casar um nome de bairro 'bruto' (como vem no BO da SSP, às vezes com
    variações/sufixos) com o nome oficial usado em POPULACAO_BAIRROS_POA."""
    if not bruto_norm:
        return None
    if bruto_norm in oficiais_norm:
        return bruto_norm
    # alias explícito (variações frequentes mapeadas manualmente)
    alias = _ALIASES_BAIRRO_POA.get(bruto_norm)
    if alias:
        alias_norm = normalizar(alias)
        if alias_norm in oficiais_norm:
            return alias_norm
    # match por contenção (ex: "CENTRO HISTORICO REGIAO 1" contém "CENTRO HISTORICO")
    candidatos = [o for o in oficiais_lista if o in bruto_norm or bruto_norm in o]
    if candidatos:
        return max(candidatos, key=len)
    return None


def _baixar_ocorrencias_por_bairro_poa(headless: bool = True) -> dict:
    """Baixa os ZIPs de ocorrências desagregadas mais recentes, filtra as cidades
    da RMPA que têm dados por bairro no IBGE e conta ocorrências por (cidade, bairro)."""
    zips = _descobrir_zips_ocorrencias_desagregadas(headless=headless)
    if not zips:
        log.warning("Não achei arquivos de ocorrências desagregadas na página da SSP-RS.")
        return {}

    # conjuntos de bairros por cidade para lookup rápido
    oficiais_por_cidade = {
        cidade: set(bairros.keys())
        for cidade, bairros in POPULACAO_BAIRROS_RMPA.items()
    }
    lista_por_cidade = {
        cidade: list(bairros.keys())
        for cidade, bairros in POPULACAO_BAIRROS_RMPA.items()
    }
    cidades_com_dados = set(POPULACAO_BAIRROS_RMPA.keys())

    # contagem: {(cidade_norm, bairro_norm): n_ocorrencias}
    contagem: dict = {}
    total_linhas_rmpa = 0
    nao_reconhecidos: dict = {}
    for item in zips:
        try:
            resp = requests.get(item["href"], headers=HEADERS, timeout=120)
            resp.raise_for_status()
            conteudo_csv = _extrair_csv_de_zip(resp.content)
            if not conteudo_csv:
                log.warning("Nenhum CSV encontrado dentro do ZIP: %s", item["href"])
                continue
        except Exception as e:
            log.warning("Falha ao baixar/extrair %s: %s", item["href"], e)
            continue

        texto = conteudo_csv.decode("latin-1", errors="replace")
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(texto[:2000], delimiters=";,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        leitor = csv.DictReader(io.StringIO(texto), dialect=dialect)

        colunas = leitor.fieldnames or []
        col_municipio = next((c for c in colunas if c and "MUNIC" in c.upper()), None)
        col_bairro = next((c for c in colunas if c and "BAIRRO" in c.upper()), None)
        if not col_municipio or not col_bairro:
            log.warning("Não achei colunas de município/bairro em %s (colunas: %s)", item["href"], colunas)
            continue

        for linha in leitor:
            municipio = normalizar(str(linha.get(col_municipio, "")))
            if municipio not in cidades_com_dados:
                continue
            total_linhas_rmpa += 1
            bairro_bruto = normalizar(str(linha.get(col_bairro, "")).strip())
            if not bairro_bruto:
                continue
            resolvido = _resolver_bairro_oficial(
                bairro_bruto, oficiais_por_cidade[municipio], lista_por_cidade[municipio]
            )
            if resolvido:
                chave = (municipio, resolvido)
                contagem[chave] = contagem.get(chave, 0) + 1
            else:
                nao_reconhecidos[bairro_bruto] = nao_reconhecidos.get(bairro_bruto, 0) + 1

        cidades_encontradas = len({k[0] for k in contagem})
        bairros_encontrados = len(contagem)
        log.info("Processado %s: %d bairros em %d cidades reconhecidos até agora",
                 item["texto"], bairros_encontrados, cidades_encontradas)

    reconhecidas = sum(contagem.values())
    total = reconhecidas + sum(nao_reconhecidos.values())
    if total:
        log.info(
            "Match de bairro RMPA: %d/%d ocorrências reconhecidas (%.0f%%). "
            "Maiores não reconhecidas: %s",
            reconhecidas, total, 100 * reconhecidas / total,
            sorted(nao_reconhecidos.items(), key=lambda kv: -kv[1])[:5],
        )

    return contagem


# ---------------------------------------------------------------------------
# População por bairro das cidades da RMPA com dados no IBGE Censo 2022.
# Fonte: IBGE Censo Demográfico 2022, Tabela 9923 (população residente por bairro).
#
# Cidades COM dados por bairro: Porto Alegre (94), Canoas (17), Viamão (19),
#   Guaíba (17), Esteio (13), Novo Hamburgo (27), São Leopoldo (24), Eldorado do Sul (21)
# Cidades SEM dados por bairro no IBGE (ficam no nível de cidade):
#   Alvorada, Cachoeirinha, Gravataí, Sapucaia do Sul
# ---------------------------------------------------------------------------
_POPULACAO_BAIRROS_RMPA_RAW = [
  # PORTO ALEGRE
  ("PORTO ALEGRE", "RESTINGA", 62448),
  ("PORTO ALEGRE", "LOMBA DO PINHEIRO", 59200),
  ("PORTO ALEGRE", "SARANDI", 51539),
  ("PORTO ALEGRE", "MARIO QUINTANA", 44068),
  ("PORTO ALEGRE", "PARTENON", 43587),
  ("PORTO ALEGRE", "PETROPOLIS", 37613),
  ("PORTO ALEGRE", "SANTA ROSA DE LIMA", 34627),
  ("PORTO ALEGRE", "VILA NOVA", 32217),
  ("PORTO ALEGRE", "SANTA TEREZA", 31358),
  ("PORTO ALEGRE", "CENTRO HISTORICO", 30569),
  ("PORTO ALEGRE", "HIPICA", 28643),
  ("PORTO ALEGRE", "MENINO DEUS", 27961),
  ("PORTO ALEGRE", "RUBEM BERTA", 27930),
  ("PORTO ALEGRE", "CAVALHADA", 25209),
  ("PORTO ALEGRE", "CRISTAL", 24851),
  ("PORTO ALEGRE", "BOM JESUS", 24589),
  ("PORTO ALEGRE", "VILA SAO JOSE", 24011),
  ("PORTO ALEGRE", "JARDIM CARVALHO", 23405),
  ("PORTO ALEGRE", "PASSO DA AREIA", 22530),
  ("PORTO ALEGRE", "MORRO SANTANA", 21640),
  ("PORTO ALEGRE", "NONOAI", 20776),
  ("PORTO ALEGRE", "CORONEL APARICIO BORGES", 18966),
  ("PORTO ALEGRE", "VILA IPIRANGA", 18041),
  ("PORTO ALEGRE", "CAMAQUA", 17935),
  ("PORTO ALEGRE", "SANTANA", 17794),
  ("PORTO ALEGRE", "FARRAPOS", 17591),
  ("PORTO ALEGRE", "JARDIM ITU", 17565),
  ("PORTO ALEGRE", "TRISTEZA", 17201),
  ("PORTO ALEGRE", "JARDIM LEOPOLDINA", 16151),
  ("PORTO ALEGRE", "RIO BRANCO", 15710),
  ("PORTO ALEGRE", "GLORIA", 15248),
  ("PORTO ALEGRE", "CRISTO REDENTOR", 15144),
  ("PORTO ALEGRE", "PASSO DAS PEDRAS", 14435),
  ("PORTO ALEGRE", "COSTA E SILVA", 13585),
  ("PORTO ALEGRE", "IPANEMA", 13403),
  ("PORTO ALEGRE", "SANTO ANTONIO", 13105),
  ("PORTO ALEGRE", "TERESOPOLIS", 13072),
  ("PORTO ALEGRE", "CIDADE BAIXA", 13014),
  ("PORTO ALEGRE", "HUMAITA", 12744),
  ("PORTO ALEGRE", "AZENHA", 12064),
  ("PORTO ALEGRE", "BELA VISTA", 11819),
  ("PORTO ALEGRE", "VILA JARDIM", 11411),
  ("PORTO ALEGRE", "JARDIM BOTANICO", 11349),
  ("PORTO ALEGRE", "JARDIM SABARA", 11270),
  ("PORTO ALEGRE", "BELEM VELHO", 10893),
  ("PORTO ALEGRE", "SAO JOAO", 10621),
  ("PORTO ALEGRE", "VILA JOAO PESSOA", 10431),
  ("PORTO ALEGRE", "MONT'SERRAT", 10357),
  ("PORTO ALEGRE", "HIGIENOPOLIS", 10284),
  ("PORTO ALEGRE", "BOM FIM", 10160),
  ("PORTO ALEGRE", "MOINHOS DE VENTO", 9995),
  ("PORTO ALEGRE", "ABERTA DOS MORROS", 9854),
  ("PORTO ALEGRE", "BELEM NOVO", 9851),
  ("PORTO ALEGRE", "BOA VISTA", 9254),
  ("PORTO ALEGRE", "CASCATA", 9234),
  ("PORTO ALEGRE", "PONTA GROSSA", 8939),
  ("PORTO ALEGRE", "AUXILIADORA", 8909),
  ("PORTO ALEGRE", "FLORESTA", 8798),
  ("PORTO ALEGRE", "MEDIANEIRA", 8749),
  ("PORTO ALEGRE", "CAMPO NOVO", 8743),
  ("PORTO ALEGRE", "JARDIM LINDOIA", 7587),
  ("PORTO ALEGRE", "SAO SEBASTIAO", 7514),
  ("PORTO ALEGRE", "PITINGA", 7012),
  ("PORTO ALEGRE", "SAO GERALDO", 6948),
  ("PORTO ALEGRE", "INDEPENDENCIA", 6885),
  ("PORTO ALEGRE", "LAMI", 6677),
  ("PORTO ALEGRE", "PARQUE SANTA FE", 6673),
  ("PORTO ALEGRE", "JARDIM DO SALSO", 6576),
  ("PORTO ALEGRE", "ARQUIPELAGO", 6411),
  ("PORTO ALEGRE", "GUARUJA", 6145),
  ("PORTO ALEGRE", "LAGEADO", 5676),
  ("PORTO ALEGRE", "CHACARA DAS PEDRAS", 5639),
  ("PORTO ALEGRE", "CHAPEU DO SOL", 5547),
  ("PORTO ALEGRE", "ESPIRITO SANTO", 4953),
  ("PORTO ALEGRE", "SANTA CECILIA", 4640),
  ("PORTO ALEGRE", "SERRARIA", 4385),
  ("PORTO ALEGRE", "JARDIM EUROPA", 4372),
  ("PORTO ALEGRE", "TRES FIGUEIRAS", 4016),
  ("PORTO ALEGRE", "VILA ASSUNCAO", 3974),
  ("PORTO ALEGRE", "JARDIM SAO PEDRO", 3320),
  ("PORTO ALEGRE", "NAVEGANTES", 3315),
  ("PORTO ALEGRE", "SANTA MARIA GORETTI", 3035),
  ("PORTO ALEGRE", "BOA VISTA DO SUL", 2703),
  ("PORTO ALEGRE", "AGRONOMIA", 2677),
  ("PORTO ALEGRE", "JARDIM ISABEL", 2592),
  ("PORTO ALEGRE", "EXTREMA", 2360),
  ("PORTO ALEGRE", "JARDIM FLORESTA", 2228),
  ("PORTO ALEGRE", "PRAIA DE BELAS", 1522),
  ("PORTO ALEGRE", "SETIMO CEU", 1166),
  ("PORTO ALEGRE", "VILA CONCEICAO", 969),
  ("PORTO ALEGRE", "ANCHIETA", 791),
  ("PORTO ALEGRE", "FARROUPILHA", 774),
  ("PORTO ALEGRE", "SAO CAETANO", 733),
  ("PORTO ALEGRE", "PEDRA REDONDA", 570),
  # CANOAS
  ("CANOAS", "MATHIAS VELHO", 43325),
  ("CANOAS", "GUAJUVIRAS", 41282),
  ("CANOAS", "HARMONIA", 34740),
  ("CANOAS", "NITEROI", 32630),
  ("CANOAS", "ESTANCIA VELHA", 30519),
  ("CANOAS", "RIO BRANCO", 24578),
  ("CANOAS", "IGARA", 20213),
  ("CANOAS", "FATIMA", 19393),
  ("CANOAS", "SAO JOSE", 16837),
  ("CANOAS", "MATO GRANDE", 16377),
  ("CANOAS", "OLARIA", 16264),
  ("CANOAS", "CENTRO", 15698),
  ("CANOAS", "MARECHAL RONDON", 15251),
  ("CANOAS", "NOSSA SENHORA DAS GRACAS", 14968),
  ("CANOAS", "SAO LUIS", 4407),
  ("CANOAS", "BRIGADEIRA", 1143),
  ("CANOAS", "INDUSTRIAL", 32),
  # VIAMAO
  ("VIAMAO", "SANTA ISABEL", 23997),
  ("VIAMAO", "CECILIA", 21198),
  ("VIAMAO", "AUGUSTA", 17673),
  ("VIAMAO", "SANTO ONOFRE", 15526),
  ("VIAMAO", "SAO LUCAS", 11590),
  ("VIAMAO", "TARUMA", 11194),
  ("VIAMAO", "SAO TOME", 11045),
  ("VIAMAO", "ELSA", 9394),
  ("VIAMAO", "CENTRO", 6954),
  ("VIAMAO", "PASSO DO VIGARIO", 6904),
  ("VIAMAO", "KRAHE", 6653),
  ("VIAMAO", "FIUZA", 5713),
  ("VIAMAO", "BRANQUINHA", 5645),
  ("VIAMAO", "QUERENCIA", 5253),
  ("VIAMAO", "MARTINICA", 4531),
  ("VIAMAO", "VIAMOPOLIS", 3233),
  ("VIAMAO", "COCAO", 3139),
  ("VIAMAO", "ESTANCIA GRANDE", 144),
  ("VIAMAO", "PARQUE SAINT'HILAIRE", 112),
  # GUAIBA
  ("GUAIBA", "SANTA RITA", 19504),
  ("GUAIBA", "COLINA", 10662),
  ("GUAIBA", "BOM FIM", 8394),
  ("GUAIBA", "JARDIM IOLANDA", 6100),
  ("GUAIBA", "JARDIM DOS LAGOS", 5752),
  ("GUAIBA", "PASSO FUNDO", 5390),
  ("GUAIBA", "CENTRO", 5320),
  ("GUAIBA", "FLORIDA", 5041),
  ("GUAIBA", "PEDRAS BRANCAS", 4878),
  ("GUAIBA", "ERMO", 4243),
  ("GUAIBA", "COLUMBIA CITY", 4211),
  ("GUAIBA", "ALEGRIA", 3448),
  ("GUAIBA", "CORONEL NASSUCA", 2716),
  ("GUAIBA", "PARQUE 35", 2582),
  ("GUAIBA", "ALVORADA", 1951),
  ("GUAIBA", "ALTOS DA ALEGRIA", 469),
  ("GUAIBA", "CHAVES BARCELLOS", 36),
  # ESTEIO
  ("ESTEIO", "CENTRO", 10064),
  ("ESTEIO", "SAO SEBASTIAO", 8122),
  ("ESTEIO", "PARQUE PRIMAVERA", 7687),
  ("ESTEIO", "SAO JOSE", 6867),
  ("ESTEIO", "SANTO INACIO", 6683),
  ("ESTEIO", "NOVO ESTEIO", 6085),
  ("ESTEIO", "PRIMAVERA", 5642),
  ("ESTEIO", "TRES MARIAS", 5289),
  ("ESTEIO", "OLIMPICA", 5102),
  ("ESTEIO", "JARDIM PLANALTO", 4635),
  ("ESTEIO", "LIBERDADE", 4221),
  ("ESTEIO", "TAMANDARE", 4208),
  ("ESTEIO", "TRES PORTOS", 1017),
  # NOVO HAMBURGO
  ("NOVO HAMBURGO", "CANUDOS", 56453),
  ("NOVO HAMBURGO", "SANTO AFONSO", 21920),
  ("NOVO HAMBURGO", "BOA SAUDE", 12773),
  ("NOVO HAMBURGO", "RONDONIA", 11456),
  ("NOVO HAMBURGO", "SAO JORGE", 9802),
  ("NOVO HAMBURGO", "CENTRO", 8455),
  ("NOVO HAMBURGO", "DIEHL", 8052),
  ("NOVO HAMBURGO", "LOMBA GRANDE", 7716),
  ("NOVO HAMBURGO", "IDEAL", 6763),
  ("NOVO HAMBURGO", "LIBERDADE", 6553),
  ("NOVO HAMBURGO", "PRIMAVERA", 6146),
  ("NOVO HAMBURGO", "MAUA", 5937),
  ("NOVO HAMBURGO", "SAO JOSE", 5821),
  ("NOVO HAMBURGO", "VILA NOVA", 5564),
  ("NOVO HAMBURGO", "ROSELANDIA", 5296),
  ("NOVO HAMBURGO", "RINCAO", 5019),
  ("NOVO HAMBURGO", "GUARANI", 4954),
  ("NOVO HAMBURGO", "OPERARIO", 4554),
  ("NOVO HAMBURGO", "PATRIA NOVA", 4272),
  ("NOVO HAMBURGO", "RIO BRANCO", 4014),
  ("NOVO HAMBURGO", "PETROPOLIS", 3592),
  ("NOVO HAMBURGO", "INDUSTRIAL", 3441),
  ("NOVO HAMBURGO", "VILA ROSA", 3027),
  ("NOVO HAMBURGO", "OURO BRANCO", 2970),
  ("NOVO HAMBURGO", "BOA VISTA", 2943),
  ("NOVO HAMBURGO", "HAMBURGO VELHO", 2487),
  ("NOVO HAMBURGO", "ALPES DO VALE", 2142),
  # SAO LEOPOLDO
  ("SAO LEOPOLDO", "FEITORIA", 33063),
  ("SAO LEOPOLDO", "SANTOS DUMONT", 30142),
  ("SAO LEOPOLDO", "ARROIO DA MANTEIGA", 22229),
  ("SAO LEOPOLDO", "CAMPINA", 13409),
  ("SAO LEOPOLDO", "SCHARLAU", 13146),
  ("SAO LEOPOLDO", "VICENTINA", 11864),
  ("SAO LEOPOLDO", "CENTRO", 11861),
  ("SAO LEOPOLDO", "DUQUE DE CAXIAS", 11767),
  ("SAO LEOPOLDO", "CAMPESTRE", 11186),
  ("SAO LEOPOLDO", "SAO MIGUEL", 8563),
  ("SAO LEOPOLDO", "SANTO ANDRE", 6723),
  ("SAO LEOPOLDO", "SANTA TERESA", 5869),
  ("SAO LEOPOLDO", "JARDIM AMERICA", 4741),
  ("SAO LEOPOLDO", "RIO BRANCO", 4647),
  ("SAO LEOPOLDO", "PINHEIRO", 3793),
  ("SAO LEOPOLDO", "MORRO DO ESPELHO", 3657),
  ("SAO LEOPOLDO", "RIO DOS SINOS", 3582),
  ("SAO LEOPOLDO", "CRISTO REI", 3147),
  ("SAO LEOPOLDO", "FAZENDA SAO BORJA", 2758),
  ("SAO LEOPOLDO", "BOA VISTA", 2387),
  ("SAO LEOPOLDO", "SAO JOSE", 2226),
  ("SAO LEOPOLDO", "SAO JOAO BATISTA", 1633),
  ("SAO LEOPOLDO", "PADRE REUS", 1474),
  ("SAO LEOPOLDO", "FIAO", 1340),
  # ELDORADO DO SUL
  ("ELDORADO DO SUL", "CENTRO NOVO", 6197),
  ("ELDORADO DO SUL", "SANS SOUCI", 4101),
  ("ELDORADO DO SUL", "CIDADE VERDE", 3439),
  ("ELDORADO DO SUL", "MEDIANEIRA", 2856),
  ("ELDORADO DO SUL", "PARQUE ELDORADO 1", 2599),
  ("ELDORADO DO SUL", "CHACARA", 2470),
  ("ELDORADO DO SUL", "CENTRO", 2174),
  ("ELDORADO DO SUL", "RESIDENCIAL ELDORADO", 1416),
  ("ELDORADO DO SUL", "INDUSTRIAL", 1278),
  ("ELDORADO DO SUL", "PICADA", 1255),
  ("ELDORADO DO SUL", "LOTEAMENTO POPULAR", 1228),
  ("ELDORADO DO SUL", "PARQUE ELDORADO 3", 1178),
  ("ELDORADO DO SUL", "PROGRESSO", 1168),
  ("ELDORADO DO SUL", "PARQUE ELDORADO 2", 1132),
  ("ELDORADO DO SUL", "ITAI", 990),
  ("ELDORADO DO SUL", "SOL NASCENTE", 783),
  ("ELDORADO DO SUL", "VILA DA PAZ", 742),
  ("ELDORADO DO SUL", "PARQUE DAS ACACIAS SUL", 367),
  ("ELDORADO DO SUL", "BOM RETIRO", 231),
  ("ELDORADO DO SUL", "PARQUE DAS ACACIAS NORTE", 155),
  ("ELDORADO DO SUL", "PARQUE ELDORADO 4", 99),
]

# Dicionário normalizado: {cidade_norm: {bairro_norm: populacao}}
POPULACAO_BAIRROS_RMPA: dict = {}
for _cidade, _bairro, _pop in _POPULACAO_BAIRROS_RMPA_RAW:
    _cn = normalizar(_cidade)
    _bn = normalizar(_bairro)
    POPULACAO_BAIRROS_RMPA.setdefault(_cn, {})[_bn] = _pop

# Compatibilidade com código antigo
POPULACAO_BAIRROS_POA: dict = POPULACAO_BAIRROS_RMPA.get("PORTO ALEGRE", {})


def obter_risco_bairro_poa(cache_path: Path, headless: bool = True) -> dict:
    """Retorna {cidade_norm: {bairro_norm: {ocorrencias, populacao, taxa_100k, risco}}}.

    Agora cobre POA + Canoas + Viamão + Guaíba + Esteio + Novo Hamburgo +
    São Leopoldo + Eldorado do Sul (todas cidades da RMPA com dados por bairro no IBGE).
    Cidades sem dados por bairro (Alvorada, Cachoeirinha, Gravataí, Sapucaia)
    continuam usando o nível de município.
    O risco é calculado por tercis DENTRO de cada cidade, para ser comparável
    entre bairros da mesma localidade."""
    import json

    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text(encoding="utf-8"))
            idade_dias = (datetime.now() - datetime.fromisoformat(cache["gerado_em"])).days
            if idade_dias < CACHE_RISCO_DIAS:
                log.info("Usando cache de criminalidade por bairro RMPA (%d dia(s) de idade).", idade_dias)
                return cache["bairros"]
        except Exception as e:
            log.warning("Cache de risco por bairro inválido, recalculando: %s", e)

    if not POPULACAO_BAIRROS_RMPA:
        return {}

    ocorrencias_por_chave = _baixar_ocorrencias_por_bairro_poa(headless=headless)
    if not ocorrencias_por_chave:
        return {}

    # resultado: {cidade_norm: {bairro_norm: {...}}}
    resultado: dict = {}
    for cidade, bairros_pop in POPULACAO_BAIRROS_RMPA.items():
        resultado[cidade] = {}
        for bairro, populacao in bairros_pop.items():
            ocorrencias = ocorrencias_por_chave.get((cidade, bairro), 0)
            taxa = (ocorrencias / populacao) * 100_000 if populacao else 0
            resultado[cidade][bairro] = {
                "ocorrencias": ocorrencias,
                "populacao": populacao,
                "taxa_100k": round(taxa, 1),
            }
        # tercis calculados dentro da cidade (comparar bairros entre si)
        taxas_cidade = [v["taxa_100k"] for v in resultado[cidade].values()]
        for bairro in resultado[cidade]:
            resultado[cidade][bairro]["risco"] = _classificar_risco(
                resultado[cidade][bairro]["taxa_100k"], taxas_cidade
            )
        n_bairros = len(resultado[cidade])
        n_com_dados = sum(1 for v in resultado[cidade].values() if v["ocorrencias"] > 0)
        log.info("%s: %d/%d bairros com ocorrências registradas", cidade, n_com_dados, n_bairros)

    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(
            json.dumps({"gerado_em": datetime.now().isoformat(), "bairros": resultado}, ensure_ascii=False),
            encoding="utf-8",
        )
    except OSError as e:
        log.warning("Não consegui salvar cache de risco por bairro: %s", e)

    return resultado


def explorar_dados_abertos_ssp(headless: bool = True, tentativas: int = 3) -> list:
    """Reconhecimento: abre https://www.ssp.rs.gov.br/dados-abertos com um Chrome
    de verdade e lista os links de download (csv/xlsx/xls/zip) que a página tem.

    Essa página tem os microdados desagregados (com campo de bairro, Lei
    15.610/2021), mas robots.txt bloqueia fetch automatizado simples -- por
    isso usamos o navegador real (mesma técnica do CSV da Caixa). Tem retry
    porque o site já se mostrou instável (ERR_CONNECTION_RESET esporádico)."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as e:
        raise RuntimeError("Playwright não instalado. Rode: pip install playwright && playwright install chromium") from e

    url = "https://www.ssp.rs.gov.br/dados-abertos"
    encontrados = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        contexto = browser.new_context(user_agent=HEADERS["User-Agent"], locale="pt-BR")
        pagina = contexto.new_page()

        for tentativa in range(1, tentativas + 1):
            try:
                pagina.goto(url, wait_until="networkidle", timeout=45000)
                pagina.wait_for_timeout(2000)
                break
            except Exception as e:
                log.warning("Tentativa %d/%d de abrir %s falhou: %s", tentativa, tentativas, url, e)
                if tentativa == tentativas:
                    browser.close()
                    raise
                pagina.wait_for_timeout(5000 * tentativa)  # espera crescente entre tentativas

        links = pagina.eval_on_selector_all(
            "a[href]",
            "els => els.map(e => ({href: e.href, texto: e.textContent.trim()}))",
        )
        for link in links:
            href = link["href"].lower()
            if any(href.endswith(ext) for ext in (".csv", ".xlsx", ".xls", ".zip")):
                encontrados.append(link)

        browser.close()

    log.info("Encontrados %d links de download em %s:", len(encontrados), url)
    for link in encontrados:
        log.info("  %s -> %s", link["texto"], link["href"])

    return encontrados


def processar_uf(
    uf: str,
    outdir: Path,
    cidades: list[str] | None = None,
    html_out: Path | None = None,
    metodo: str = "navegador",
    headless: bool = True,
    com_risco: bool = True,
):
    if metodo == "navegador":
        conteudo = baixar_csv_via_navegador(uf, headless=headless)
    else:
        conteudo = baixar_csv(uf)

    if parece_desafio_radware(conteudo):
        log.error(
            "%s: recebi a página de desafio do Radware em vez do CSV. "
            "O bloqueio não foi resolvido -- ver instruções de troubleshooting.",
            uf,
        )
        return

    linhas = parsear_csv(conteudo)
    if not linhas:
        log.warning("Nenhuma linha encontrada para %s (site pode ter mudado o formato).", uf)
        # ainda assim gera um HTML vazio, pra não quebrar o commit do workflow
        if html_out:
            gerar_html(uf, datetime.now().strftime("%Y-%m-%d"), [], [], [], cidades, html_out)
        return

    if cidades:
        linhas_filtradas = filtrar_por_cidade(linhas, cidades)
        if not linhas_filtradas:
            log.warning(
                "Filtro de cidade não retornou nada -- confira se os nomes em "
                "CIDADES_RMPA batem com o texto real do CSV (rode sem --cidades pra conferir)."
            )
        linhas = linhas_filtradas

    # --- Adiciona 'Origem' aos imóveis da Caixa ---
    for row in linhas:
        if "Origem" not in row:
            row["Origem"] = "Caixa"

    # --- Integração Mega Leilões ---
    try:
        log.info("Iniciando scraper do Mega Leilões...")
        ml_scraper = MegaLeiloesScraper(cidades_alvo=cidades if cidades else [])
        imoveis_ml = ml_scraper.raspar()
        
        # Pega as chaves reais do CSV da Caixa para não dar erro no csv.DictWriter
        chaves_caixa = list(linhas[0].keys()) if linhas else []
        id_col = achar_coluna_id(linhas) if linhas else "N° do imóvel"
        
        for imovel in imoveis_ml:
            novo_row = {k: "" for k in chaves_caixa}
            
            # Preenche os campos principais
            if id_col in novo_row: novo_row[id_col] = imovel["id"]
            if "Cidade" in novo_row: novo_row["Cidade"] = imovel["cidade"]
            if "Bairro" in novo_row: novo_row["Bairro"] = imovel["bairro"]
            if "Endereço" in novo_row: novo_row["Endereço"] = imovel["endereco"]
            if "Preço" in novo_row: novo_row["Preço"] = str(imovel["preco"]).replace(".", ",")
            if "Valor de avaliação" in novo_row: novo_row["Valor de avaliação"] = str(imovel["avaliacao"]).replace(".", ",")
            if "Desconto" in novo_row: novo_row["Desconto"] = str(imovel["desconto"]).replace(".", ",")
            if "Modalidade de venda" in novo_row: novo_row["Modalidade de venda"] = imovel["modalidade"]
            if "Descrição" in novo_row: novo_row["Descrição"] = imovel["descricao"]
            if "Link de acesso" in novo_row: novo_row["Link de acesso"] = imovel["link"]
            if "Origem" in novo_row: novo_row["Origem"] = "Mega Leilões"
            
            linhas.append(novo_row)
            
        log.info("Adicionados %d imóveis do Mega Leilões", len(imoveis_ml))
    except Exception as e:
        log.error("Erro ao rodar scraper do Mega Leilões: %s", e)
    # --- Integração Leilões Judiciais ---
    try:
        log.info("Iniciando scraper do Leilões Judiciais...")
        lj_scraper = LeiloesJudiciaisScraper(cidades_alvo=cidades if cidades else [])
        imoveis_lj = lj_scraper.raspar()
        
        for imovel in imoveis_lj:
            novo_row = {k: "" for k in chaves_caixa}
            if id_col in novo_row: novo_row[id_col] = imovel["id"]
            if "Cidade" in novo_row: novo_row["Cidade"] = imovel["cidade"]
            if "Bairro" in novo_row: novo_row["Bairro"] = imovel["bairro"]
            if "Endereço" in novo_row: novo_row["Endereço"] = imovel["endereco"]
            if "Preço" in novo_row: novo_row["Preço"] = str(imovel["preco"]).replace(".", ",")
            if "Valor de avaliação" in novo_row: novo_row["Valor de avaliação"] = str(imovel["avaliacao"]).replace(".", ",")
            if "Desconto" in novo_row: novo_row["Desconto"] = f'{imovel["desconto"]:.2f}'.replace(".", ",")
            if "Modalidade de venda" in novo_row: novo_row["Modalidade de venda"] = imovel["modalidade"]
            if "Descrição" in novo_row: novo_row["Descrição"] = imovel["descricao"]
            if "Link de acesso" in novo_row: novo_row["Link de acesso"] = imovel["link"]
            if "Origem" in novo_row: novo_row["Origem"] = "Leilões Judiciais"
            linhas.append(novo_row)
            
        log.info("Adicionados %d imóveis do Leilões Judiciais", len(imoveis_lj))
    except Exception as e:
        log.error("Erro ao rodar scraper do Leilões Judiciais: %s", e)
    # -------------------------------

    id_col = achar_coluna_id(linhas) if linhas else ""
    data_atual = datetime.now().strftime("%Y-%m-%d")
    anteriores = carregar_snapshot_anterior(uf, outdir, data_atual)
    salvar_snapshot(linhas, uf, outdir)

    novos, removidos = comparar(linhas, anteriores, id_col) if linhas else ([], [])
    log.info("%s: %d novos, %d removidos (total atual: %d)", uf, len(novos), len(removidos), len(linhas))

    risco_por_municipio = {}
    risco_bairro_poa = {}
    if com_risco:
        try:
            risco_por_municipio = obter_risco_por_municipio(outdir / "cache_criminalidade.json")
        except Exception as e:
            log.warning("Falha ao obter índice de risco por município, seguindo sem essa coluna: %s", e)
        try:
            risco_bairro_poa = obter_risco_bairro_poa(outdir / "cache_criminalidade_bairro_poa.json", headless=headless)
        except Exception as e:
            log.warning("Falha ao obter índice de risco por bairro (POA): %s", e)

    import json
    historico_path = outdir / "historico_precos.json"
    historico_precos = {}
    if historico_path.exists():
        try:
            historico_precos = json.loads(historico_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    if linhas:
        for row in linhas:
            imovel_id = row.get(id_col)
            if not imovel_id: continue
            preco_atual = _parar_valor_brl(row.get("Preço", ""))
            
            if imovel_id not in historico_precos:
                historico_precos[imovel_id] = {"datas": [data_atual], "precos": [preco_atual]}
            else:
                ultimo_preco = historico_precos[imovel_id]["precos"][-1]
                if abs(ultimo_preco - preco_atual) > 0.01:
                    historico_precos[imovel_id]["datas"].append(data_atual)
                    historico_precos[imovel_id]["precos"].append(preco_atual)
                    
        historico_path.write_text(json.dumps(historico_precos, indent=2, ensure_ascii=False), encoding="utf-8")

    if html_out:
        gerar_html(uf, data_atual, novos, removidos, linhas, cidades, html_out, risco_por_municipio, risco_bairro_poa, historico_precos)


def main():
    parser = argparse.ArgumentParser(description="Scraper diário de leilões Caixa")
    parser.add_argument("--ufs", nargs="+", default=["RS"], help="Ex: RS SP RJ (default: RS)")
    parser.add_argument("--outdir", default="./data", help="Diretório de snapshots")
    parser.add_argument(
        "--cidades",
        nargs="*",
        default=CIDADES_RMPA,
        help="Cidades para filtrar (default: Região Metropolitana de Porto Alegre). "
        "Passe --cidades sem valores pra desativar o filtro.",
    )
    parser.add_argument(
        "--html-out",
        default="index.html",
        help="Caminho do relatório HTML gerado (default: index.html na raiz)",
    )
    parser.add_argument(
        "--open",
        action="store_true",
        help="Abre o relatório HTML automaticamente no navegador padrão ao terminar",
    )
    parser.add_argument(
        "--metodo",
        choices=["navegador", "requests"],
        default="navegador",
        help="'navegador' usa Playwright pra passar do bloqueio Radware (default, recomendado). "
        "'requests' é o método simples antigo, que costuma ser bloqueado.",
    )
    parser.add_argument(
        "--mostrar-navegador",
        action="store_true",
        help="Mostra a janela do Chrome durante o download (útil se o modo headless continuar bloqueado)",
    )
    parser.add_argument(
        "--sem-risco",
        action="store_true",
        help="Desativa a coluna de índice de risco (criminalidade por cidade, via SSP-RS)",
    )
    parser.add_argument(
        "--explorar-bairro-poa",
        action="store_true",
        help="Não roda o scraper -- só abre a página de dados abertos da SSP-RS e lista "
        "os links de download disponíveis (reconhecimento pro índice de risco por bairro)",
    )
    args = parser.parse_args()

    if args.explorar_bairro_poa:
        explorar_dados_abertos_ssp(headless=not args.mostrar_navegador)
        return

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)  # garante que a pasta existe mesmo se tudo falhar
    cidades = args.cidades if args.cidades else None
    html_out = Path(args.html_out) if args.html_out else None

    sucesso = 0
    for uf in args.ufs:
        try:
            processar_uf(
                uf,
                outdir,
                cidades=cidades,
                html_out=html_out,
                metodo=args.metodo,
                headless=not args.mostrar_navegador,
                com_risco=not args.sem_risco,
            )
            sucesso += 1
        except requests.RequestException as e:
            log.error("Erro ao baixar/processar %s: %s", uf, e)

    if sucesso == 0:
        log.error("Nenhuma UF processada com sucesso -- abortando com erro.")
        raise SystemExit(1)

    if args.open and html_out and html_out.exists():
        webbrowser.open(html_out.resolve().as_uri())


if __name__ == "__main__":
    main()
